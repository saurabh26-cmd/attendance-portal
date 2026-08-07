import os, sqlite3, calendar, io, csv, secrets, time, re
from datetime import date
from functools import wraps
from flask import Flask, request, redirect, session, send_file, render_template_string, abort
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=1800,
)
DB = os.environ.get("DB_FILE", "attendance.db")
ADMIN_LOGIN = os.environ.get("ADMIN_LOGIN", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")
LOGIN_WINDOW, MAX_LOGIN_ATTEMPTS = 300, 8
login_attempts = {}


def db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c


def add_column(c, table, column, definition):
    cols = {r[1] for r in c.execute(f"PRAGMA table_info({table})")}
    if column not in cols:
        c.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init():
    c = db()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS users(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      name TEXT NOT NULL, login_id TEXT UNIQUE NOT NULL,
      password_hash TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'manager',
      location TEXT DEFAULT '', active INTEGER DEFAULT 1,
      mobile TEXT DEFAULT '', email TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS employees(
      id INTEGER PRIMARY KEY AUTOINCREMENT, employee_code TEXT NOT NULL,
      name TEXT NOT NULL, designation TEXT NOT NULL DEFAULT 'Employee',
      location TEXT NOT NULL DEFAULT '', active INTEGER DEFAULT 1,
      mobile TEXT DEFAULT '', email TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS attendance(
      employee_id INTEGER NOT NULL, att_date TEXT NOT NULL, status TEXT NOT NULL,
      PRIMARY KEY(employee_id,att_date));
    """)
    # Migrate the existing database created by the earlier version.
    add_column(c, "users", "mobile", "TEXT DEFAULT ''")
    add_column(c, "users", "email", "TEXT DEFAULT ''")
    add_column(c, "employees", "mobile", "TEXT DEFAULT ''")
    add_column(c, "employees", "email", "TEXT DEFAULT ''")
    add_column(c, "employees", "designation", "TEXT DEFAULT 'Employee'")
    add_column(c, "employees", "location", "TEXT DEFAULT ''")

    admin = c.execute("SELECT id FROM users WHERE login_id=?", (ADMIN_LOGIN,)).fetchone()
    if not admin:
        if not ADMIN_PASSWORD:
            c.close()
            raise RuntimeError("ADMIN_PASSWORD is not configured in Render Environment Variables.")
        c.execute("""INSERT INTO users(name,login_id,password_hash,role,location,active)
                     VALUES(?,?,?,?,?,1)""",
                  ("Administrator", ADMIN_LOGIN, generate_password_hash(ADMIN_PASSWORD), "admin", "ALL"))
    c.commit(); c.close()


CSS = """<style>
*{box-sizing:border-box}body{margin:0;font-family:Arial,sans-serif;background:#f3f5f9;color:#172033}
.nav{background:#111827;color:white;padding:15px 22px;display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap}
.nav a{color:white;text-decoration:none;margin-left:12px;font-size:14px}.wrap{max-width:1450px;margin:24px auto;padding:0 16px}
.card{background:white;border:1px solid #e5e7eb;border-radius:12px;padding:18px;margin:14px 0;box-shadow:0 2px 10px #0000000b}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}input,select,button{padding:10px;border:1px solid #aeb7c4;border-radius:7px;font:inherit;width:100%;background:#fff;color:#172033}
select{appearance:auto;-webkit-appearance:auto}select option{background:#fff;color:#172033;padding:8px}button,.btn{background:#111827;color:white;border:0;cursor:pointer;padding:10px 14px;border-radius:7px;text-decoration:none;width:auto;display:inline-block}
.btn.success{background:#087f5b}.btn.secondary{background:#e9edf3;color:#172033}.btn.danger{background:#b42318}.btn.smallbtn{padding:7px 10px;font-size:12px}
table{border-collapse:collapse;width:100%;font-size:13px}th,td{border:1px solid #e2e6ec;padding:7px;text-align:center;white-space:nowrap}th{background:#f7f8fa}
.scroll{overflow:auto}.emp{text-align:left;position:sticky;left:0;background:white;z-index:2}.small{font-size:12px;color:#667085}.login{max-width:420px;margin:70px auto}
.hero{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}.actions{display:flex;gap:8px;flex-wrap:wrap}.flash{background:#e7f7ef;padding:11px;border-radius:7px;margin-bottom:10px}
.badge{padding:4px 8px;border-radius:999px;background:#eef2f7;font-size:11px}.help{background:#f8fafc;border:1px dashed #b9c2cf;padding:12px;border-radius:8px}.danger-text{color:#b42318}
</style>"""


def csrf_token():
    token = session.get("_csrf")
    if not token:
        token = secrets.token_urlsafe(32); session["_csrf"] = token
    return token


def check_csrf():
    if request.method == "POST":
        if request.form.get("_csrf", "") != session.get("_csrf", ""):
            abort(400, description="Invalid security token. Refresh the page and try again.")


def html(title, body):
    nav = """<div class='nav'><b>📋 Monthly Attendance Portal</b><div>
    {% if session.get('uid') %}<span>{{session.get('name')}} · {{session.get('role')|title}}</span>
    <a href='/dashboard'>Dashboard</a>
    {% if session.get('role')=='admin' %}<a href='/users'>Managers</a><a href='/employees'>Employees</a><a href='/settings'>Security</a>{% endif %}
    <a href='/monthly'>Attendance</a><a href='/report'>Reports</a>
    <form method='post' action='/logout' style='display:inline'><input type='hidden' name='_csrf' value='{{ token }}'><button style='background:none;padding:0;margin-left:12px;color:white'>Logout</button></form>
    {% endif %}</div></div>"""
    return f"<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{title}</title>{CSS}</head><body>{nav}<div class='wrap'>{body}</div></body></html>"


def flash(msg): session.setdefault("flash", []).append(msg)


def page(title, body, **context):
    msgs = session.pop("flash", [])
    body = body.replace("{{FLASH}}", "".join(f"<div class='flash'>{m}</div>" for m in msgs))
    context.setdefault("token", csrf_token())
    return render_template_string(html(title, body), **context)


def login_required(f):
    @wraps(f)
    def w(*a, **k):
        if not session.get("uid"): return redirect("/")
        if time.time() - session.get("_last_seen", time.time()) > 1800:
            session.clear(); return redirect("/")
        session["_last_seen"] = time.time()
        c=db(); u=c.execute("SELECT active FROM users WHERE id=?",(session["uid"],)).fetchone(); c.close()
        if not u or not u["active"]: session.clear(); return redirect("/")
        return f(*a, **k)
    return w


def admin_required(f):
    @wraps(f)
    def w(*a, **k):
        if session.get("role") != "admin": flash("Admin access required."); return redirect("/dashboard")
        return f(*a, **k)
    return w


def allowed_locations(c):
    if session.get("role") == "admin":
        return [r["location"] for r in c.execute("SELECT DISTINCT location FROM employees WHERE location<>'' ORDER BY location")]
    return [session.get("location")] if session.get("location") else []


def valid_email(email): return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def parse_upload(file):
    """Return list of dicts from CSV or XLSX. XLSX needs openpyxl."""
    if not file or not file.filename: raise ValueError("Please choose a file.")
    name=file.filename.lower()
    if name.endswith(".csv"):
        text=file.read().decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if name.endswith(".xlsx"):
        try: from openpyxl import load_workbook
        except ImportError: raise ValueError("Excel support is unavailable. Add openpyxl to requirements.txt and redeploy.")
        wb=load_workbook(file, read_only=True, data_only=True); ws=wb.active
        rows=list(ws.iter_rows(values_only=True)); wb.close()
        if not rows: return []
        headers=[str(x).strip() if x is not None else "" for x in rows[0]]
        return [dict(zip(headers,["" if x is None else str(x).strip() for x in r])) for r in rows[1:] if any(x is not None for x in r)]
    raise ValueError("Only .xlsx or .csv files are supported.")


def template_xlsx(headers, filename):
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title="Upload"
    ws.append(headers)
    for cell in ws[1]: cell.font=__import__('openpyxl').styles.Font(bold=True)
    ws.freeze_panes="A2"
    for i,h in enumerate(headers,1): ws.column_dimensions[chr(64+i)].width=max(16,len(h)+2)
    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio,as_attachment=True,download_name=filename,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.before_request
def security_checks():
    if session.get("uid") and time.time()-session.get("_last_seen",time.time())>1800:
        session.clear(); return redirect("/")
    if session.get("uid"): session["_last_seen"]=time.time()


@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"]="nosniff"; response.headers["X-Frame-Options"]="DENY"
    response.headers["Referrer-Policy"]="strict-origin-when-cross-origin"; response.headers["Cache-Control"]="no-store"
    response.headers["Permissions-Policy"]="camera=(), microphone=(), geolocation=()"
    response.headers["Content-Security-Policy"]="default-src 'self'; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; img-src 'self' data:; font-src 'self' data:; frame-ancestors 'none'; base-uri 'self'; form-action 'self'"
    return response


@app.route('/',methods=['GET','POST'])
def login():
    if session.get('uid'): return redirect('/dashboard')
    if request.method=='POST':
        check_csrf(); key=request.remote_addr or 'unknown'; now=time.time(); recent=[t for t in login_attempts.get(key,[]) if now-t<LOGIN_WINDOW]
        if len(recent)>=MAX_LOGIN_ATTEMPTS: flash('Too many failed attempts. Try again later.'); return redirect('/')
        login_id=request.form.get('login_id','').strip(); password=request.form.get('password',''); c=db(); u=c.execute('SELECT * FROM users WHERE login_id=? AND active=1',(login_id,)).fetchone(); c.close()
        if u and check_password_hash(u['password_hash'],password):
            login_attempts.pop(key,None); session.clear(); session.permanent=True; session.update(uid=u['id'],name=u['name'],role=u['role'],location=u['location']); csrf_token(); return redirect('/dashboard')
        recent.append(now); login_attempts[key]=recent; flash('Invalid Login ID or Password.')
    token=csrf_token()
    return page('Login',"""{{FLASH}}<div class='card login'><h2>Welcome 👋</h2><p class='small'>Monthly attendance management system</p>
    <form method='post'><input type='hidden' name='_csrf' value='{{ token }}'><label>Login ID<input name='login_id' autocomplete='username' required></label><br>
    <label>Password<input type='password' name='password' autocomplete='current-password' required></label><br><button>Login</button></form></div>""",token=token)


@app.route('/logout',methods=['POST'])
@login_required
def logout(): check_csrf(); session.clear(); return redirect('/')


@app.route('/dashboard')
@login_required
def dashboard():
    c=db(); emp=c.execute("SELECT COUNT(*) n FROM employees WHERE active=1" + (" AND location=?" if session['role']!='admin' else ""), ((session['location'],) if session['role']!='admin' else ())).fetchone()['n']
    mgr=c.execute("SELECT COUNT(*) n FROM users WHERE role='manager' AND active=1").fetchone()['n']; c.close()
    loc='All Locations' if session['role']=='admin' else session['location']
    return page('Dashboard',f"""{{{{FLASH}}}}<div class='hero'><div><h1>Dashboard</h1><p class='small'>Welcome, {session['name']}.</p></div><a class='btn' href='/monthly'>➕ Fill Attendance</a></div>
    <div class='grid'><div class='card'><span class='small'>Active Employees</span><h2>{emp}</h2></div><div class='card'><span class='small'>Managers</span><h2>{mgr}</h2></div><div class='card'><span class='small'>Location</span><h3>{loc}</h3></div></div>""")


@app.route('/settings',methods=['GET','POST'])
@login_required
@admin_required
def settings():
    if request.method=='POST':
        check_csrf(); current=request.form.get('current_password',''); new=request.form.get('new_password',''); confirm=request.form.get('confirm_password','')
        c=db(); u=c.execute("SELECT * FROM users WHERE id=?",(session['uid'],)).fetchone()
        if not check_password_hash(u['password_hash'],current): flash('Current admin password is incorrect.')
        elif len(new)<10: flash('New password must be at least 10 characters.')
        elif new!=confirm: flash('New password and confirmation do not match.')
        else:
            c.execute("UPDATE users SET password_hash=? WHERE id=?",(generate_password_hash(new),session['uid'])); c.commit(); flash('Admin password updated successfully.')
        c.close()
    return page('Security',"""{{FLASH}}<h1>Security</h1><div class='card'><h3>Change Admin Password</h3><form method='post' class='grid'><input type='hidden' name='_csrf' value='{{ token }}'>
    <input type='password' name='current_password' placeholder='Current password' required><input type='password' name='new_password' placeholder='New password (10+ characters)' required><input type='password' name='confirm_password' placeholder='Confirm new password' required><button>Update Admin Password</button></form></div>""")


@app.route('/users',methods=['GET','POST'])
@login_required
@admin_required
def users():
    c=db()
    if request.method=='POST':
        check_csrf(); action=request.form.get('action','')
        try:
            if action=='create':
                c.execute("INSERT INTO users(name,login_id,password_hash,role,location,mobile,email) VALUES(?,?,?,?,?,?,?)",(request.form['name'].strip(),request.form['login_id'].strip(),generate_password_hash(request.form['password']), 'manager',request.form['location'].strip(),request.form.get('mobile','').strip(),request.form.get('email','').strip()))
                flash('Manager created.')
            elif action=='edit':
                uid=int(request.form['id']); pw=request.form.get('password','').strip()
                if pw:
                    c.execute("UPDATE users SET name=?,login_id=?,location=?,mobile=?,email=?,password_hash=? WHERE id=? AND role='manager'",(request.form['name'].strip(),request.form['login_id'].strip(),request.form['location'].strip(),request.form.get('mobile','').strip(),request.form.get('email','').strip(),generate_password_hash(pw),uid))
                else:
                    c.execute("UPDATE users SET name=?,login_id=?,location=?,mobile=?,email=? WHERE id=? AND role='manager'",(request.form['name'].strip(),request.form['login_id'].strip(),request.form['location'].strip(),request.form.get('mobile','').strip(),request.form.get('email','').strip(),uid))
                flash('Manager updated.')
            elif action=='toggle':
                c.execute("UPDATE users SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id=? AND role='manager'",(int(request.form['id']),)); flash('Manager status updated.')
            c.commit()
        except sqlite3.IntegrityError: c.rollback(); flash('Login ID already exists.')
        except Exception as e: c.rollback(); flash('Could not save manager: '+str(e))

    locs=[r['location'] for r in c.execute("SELECT DISTINCT location FROM employees WHERE location<>'' ORDER BY location")]
    managers=c.execute("SELECT * FROM users WHERE role='manager' ORDER BY name").fetchall(); c.close()
    opts=''.join(f"<option value='{x}'>{x}</option>" for x in locs) or "<option value=''>Add employees first</option>"
    rows=''
    for u in managers:
        rows+=f"""<tr><td>{u['name']}</td><td>{u['mobile'] or ''}</td><td>{u['email'] or ''}</td><td>{u['login_id']}</td><td>{u['location']}</td><td>{'Active' if u['active'] else 'Inactive'}</td>
        <td><details><summary class='btn smallbtn'>Edit</summary><form method='post' class='grid' style='min-width:720px;text-align:left;margin-top:8px'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='hidden' name='action' value='edit'><input type='hidden' name='id' value='{u['id']}'><input name='name' value='{u['name']}' required><input name='mobile' value='{u['mobile'] or ''}'><input name='email' value='{u['email'] or ''}'><input name='login_id' value='{u['login_id']}' required><input name='password' placeholder='New password (leave blank to keep)'><select name='location'>{''.join(f"<option {'selected' if x==u['location'] else ''}>{x}</option>" for x in locs)}</select><button>Save</button></form></details>
        <form method='post' style='display:inline'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='hidden' name='action' value='toggle'><input type='hidden' name='id' value='{u['id']}'><button class='btn smallbtn danger'>Toggle</button></form></td></tr>"""
    body=f"""{{{{FLASH}}}}<div class='hero'><h1>Manager Accounts</h1><a class='btn secondary' href='/template/managers'>⬇ Download Manager Format</a></div>
    <div class='card'><h3>Add Manager</h3><form method='post' class='grid'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='hidden' name='action' value='create'><input name='name' placeholder='Name' required><input name='mobile' placeholder='Mobile Number'><input name='email' placeholder='Email ID'><input name='login_id' placeholder='Login ID' required><input type='password' name='password' placeholder='Password' required><select name='location' required>{opts}</select><button>Create Manager</button></form></div>
    <div class='card'><h3>Bulk Manager Upload</h3><p class='small'>Download the format, fill it, then upload .xlsx or .csv.</p><form method='post' action='/users/bulk' enctype='multipart/form-data'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='file' name='file' accept='.xlsx,.csv' required><button>Upload Managers</button></form></div>
    <div class='card scroll'><table><tr><th>Name</th><th>Mobile</th><th>Email</th><th>Login ID</th><th>Location</th><th>Status</th><th>Actions</th></tr>{rows}</table></div>"""
    return page('Managers',body)


@app.route('/users/bulk',methods=['POST'])
@login_required
@admin_required
def users_bulk():
    check_csrf(); c=db()
    try:
        rows=parse_upload(request.files.get('file')); required={'Name','Mobile','Email','Login ID','Password','Location'}
        if not rows: raise ValueError('The file is empty.')
        if not required.issubset(set(rows[0].keys())): raise ValueError('Use the downloaded manager format: Name, Mobile, Email, Login ID, Password, Location.')
        count=0
        for r in rows:
            vals=[r.get(k,'').strip() for k in required]
            if not r.get('Name','').strip() or not r.get('Login ID','').strip() or not r.get('Password','').strip() or not r.get('Location','').strip(): continue
            if r.get('Email','').strip() and not valid_email(r['Email'].strip()): continue
            c.execute("INSERT INTO users(name,login_id,password_hash,role,location,mobile,email) VALUES(?,?,?,?,?,?,?)",(r['Name'].strip(),r['Login ID'].strip(),generate_password_hash(r['Password'].strip()),'manager',r['Location'].strip(),r.get('Mobile','').strip(),r.get('Email','').strip())); count+=1
        c.commit(); flash(f'{count} manager(s) uploaded successfully.')
    except sqlite3.IntegrityError: c.rollback(); flash('Upload stopped: duplicate Login ID found. Fix duplicates and upload again.')
    except Exception as e: c.rollback(); flash(str(e))
    c.close(); return redirect('/users')


@app.route('/template/managers')
@login_required
@admin_required
def manager_template(): return template_xlsx(['Name','Mobile','Email','Login ID','Password','Location'],'manager_upload_format.xlsx')


@app.route('/employees',methods=['GET','POST'])
@login_required
@admin_required
def employees():
    c=db()
    if request.method=='POST':
        check_csrf(); action=request.form.get('action','')
        try:
            if action=='create':
                next_id=(c.execute("SELECT COALESCE(MAX(id),0)+1 n FROM employees").fetchone()['n'])
                code='EMP'+str(next_id).zfill(4)
                c.execute("INSERT INTO employees(employee_code,name,designation,location,mobile,email) VALUES(?,?,?,?,?,?)",(code,request.form['name'].strip(),'Employee',request.form['location'].strip(),request.form.get('mobile','').strip(),request.form.get('email','').strip()))
                flash('Employee added.')
            elif action=='edit':
                c.execute("UPDATE employees SET name=?,mobile=?,email=?,location=? WHERE id=?",(request.form['name'].strip(),request.form.get('mobile','').strip(),request.form.get('email','').strip(),request.form['location'].strip(),int(request.form['id']))); flash('Employee updated.')
            elif action=='toggle':
                c.execute("UPDATE employees SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id=?",(int(request.form['id']),)); flash('Employee status updated.')
            c.commit()
        except Exception as e: c.rollback(); flash('Could not save employee: '+str(e))
    emps=c.execute("SELECT * FROM employees ORDER BY location,name").fetchall(); c.close()
    locs=sorted({e['location'] for e in emps if e['location']})
    opts=''.join(f"<option value='{x}'>{x}</option>" for x in locs) or "<option value=''>Add a location</option>"
    rows=''
    for e in emps:
        rows+=f"""<tr><td>{e['employee_code']}</td><td>{e['name']}</td><td>{e['mobile'] or ''}</td><td>{e['email'] or ''}</td><td>{e['location']}</td><td>{'Active' if e['active'] else 'Inactive'}</td><td><details><summary class='btn smallbtn'>Edit</summary><form method='post' class='grid' style='min-width:600px;text-align:left;margin-top:8px'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='hidden' name='action' value='edit'><input type='hidden' name='id' value='{e['id']}'><input name='name' value='{e['name']}' required><input name='mobile' value='{e['mobile'] or ''}'><input name='email' value='{e['email'] or ''}'><input name='location' value='{e['location']}' required><button>Save</button></form></details><form method='post' style='display:inline'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='hidden' name='action' value='toggle'><input type='hidden' name='id' value='{e['id']}'><button class='btn smallbtn danger'>Toggle</button></form></td></tr>"""
    body=f"""{{{{FLASH}}}}<div class='hero'><h1>Employee Master</h1><a class='btn secondary' href='/template/employees'>⬇ Download Employee Format</a></div>
    <div class='card'><h3>Add Employee</h3><form method='post' class='grid'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input type='hidden' name='action' value='create'><input name='name' placeholder='Name' required><input name='mobile' placeholder='Mobile Number'><input name='email' placeholder='Email ID'><input name='location' placeholder='Warehouse / Location' required><button>Add Employee</button></form></div>
    <div class='card'><h3>Bulk Employee Upload</h3><p class='small'>Only Name, Mobile and Email are required in the file. Select the common warehouse/location below for all uploaded employees.</p><form method='post' action='/employees/bulk' enctype='multipart/form-data' class='grid'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><input name='location' placeholder='Warehouse / Location for all rows' required><input type='file' name='file' accept='.xlsx,.csv' required><button>Upload Employees</button></form></div>
    <div class='card scroll'><table><tr><th>Employee ID</th><th>Name</th><th>Mobile</th><th>Email</th><th>Location</th><th>Status</th><th>Actions</th></tr>{rows}</table></div>"""
    return page('Employees',body)


@app.route('/employees/bulk',methods=['POST'])
@login_required
@admin_required
def employees_bulk():
    check_csrf(); location=request.form.get('location','').strip(); c=db()
    try:
        if not location: raise ValueError('Please enter a warehouse/location.')
        rows=parse_upload(request.files.get('file')); required={'Name','Mobile','Email'}
        if not rows: raise ValueError('The file is empty.')
        if not required.issubset(set(rows[0].keys())): raise ValueError('Use the downloaded employee format: Name, Mobile, Email.')
        count=0; next_id=c.execute("SELECT COALESCE(MAX(id),0)+1 n FROM employees").fetchone()['n']
        for r in rows:
            name=r.get('Name','').strip(); email=r.get('Email','').strip(); mobile=r.get('Mobile','').strip()
            if not name: continue
            if email and not valid_email(email): continue
            code='EMP'+str(next_id).zfill(4); next_id+=1
            c.execute("INSERT INTO employees(employee_code,name,designation,location,mobile,email) VALUES(?,?,?,?,?,?)",(code,name,'Employee',location,mobile,email)); count+=1
        c.commit(); flash(f'{count} employee(s) uploaded successfully to {location}.')
    except Exception as e: c.rollback(); flash(str(e))
    c.close(); return redirect('/employees')


@app.route('/template/employees')
@login_required
@admin_required
def employee_template(): return template_xlsx(['Name','Mobile','Email'],'employee_upload_format.xlsx')


@app.route('/monthly',methods=['GET','POST'])
@login_required
def monthly():
    if request.method=='POST': check_csrf()
    c=db(); locs=allowed_locations(c)
    if not locs: c.close(); return page('Attendance',"<div class='card'><h2>No locations available</h2><p>Add employees first.</p></div>")
    location=request.args.get('location') or locs[0]
    if location not in locs: location=locs[0]
    year=int(request.args.get('year',date.today().year)); month=int(request.args.get('month',date.today().month)); month=max(1,min(12,month)); days=calendar.monthrange(year,month)[1]
    emps=c.execute("SELECT * FROM employees WHERE active=1 AND location=? ORDER BY name",(location,)).fetchall()
    if request.method=='POST':
        for e in emps:
            for d in range(1,days+1):
                s=request.form.get(f"a_{e['id']}_{d}",'')
                if s: c.execute("INSERT INTO attendance(employee_id,att_date,status) VALUES(?,?,?) ON CONFLICT(employee_id,att_date) DO UPDATE SET status=excluded.status",(e['id'],f'{year:04d}-{month:02d}-{d:02d}',s))
        c.commit(); c.close(); flash(f'Attendance saved for {calendar.month_name[month]} {year}.'); return redirect(f'/monthly?location={location}&year={year}&month={month}')
    start=f'{year:04d}-{month:02d}-01'; end=f'{year:04d}-{month:02d}-{days:02d}'; data={}
    for e in emps: data[e['id']]={r['att_date']:r['status'] for r in c.execute("SELECT att_date,status FROM attendance WHERE employee_id=? AND att_date BETWEEN ? AND ?",(e['id'],start,end))}
    c.close(); opts_loc=''.join(f"<option value='{x}' {'selected' if x==location else ''}>{x}</option>" for x in locs); heads=''.join(f'<th>{d}</th>' for d in range(1,days+1)); rows=''
    for e in emps:
        cells=''
        for d in range(1,days+1):
            cur=data[e['id']].get(f'{year:04d}-{month:02d}-{d:02d}',''); opts="<option value=''>-</option>"+''.join(f"<option value='{s}' {'selected' if cur==s else ''}>{s}</option>" for s in ['P','A','L','WO']); cells+=f"<td><select name='a_{e['id']}_{d}' style='width:62px'>{opts}</select></td>"
        rows+=f"<tr><td class='emp'><b>{e['employee_code']}</b><br>{e['name']}<br><span class='small'>{e['designation']}</span></td>{cells}</tr>"
    body=f"""{{{{FLASH}}}}<div class='hero'><div><h1>Monthly Attendance</h1><p class='small'>{calendar.month_name[month]} {year} · {location}</p></div><div class='actions'><a class='btn secondary' href='/report?location={location}&year={year}&month={month}'>Report</a><a class='btn success' href='/export?location={location}&year={year}&month={month}'>Export CSV</a></div></div>
    <form method='get' class='card grid'><label>Warehouse<select name='location'>{opts_loc}</select></label><label>Year<input type='number' name='year' value='{year}'></label><label>Month<input type='number' name='month' value='{month}' min='1' max='12'></label><button>Load Month</button></form>
    <form method='post'><input type='hidden' name='_csrf' value='{{{{ token }}}}'><div class='card scroll'><table><tr><th class='emp'>Employee</th>{heads}</tr>{rows}</table></div><button class='btn success'>💾 Save Full Month Attendance</button></form><p class='small'>P = Present · A = Absent · L = Leave · WO = Weekly Off</p>"""
    return page('Monthly Attendance',body)


@app.route('/report')
@login_required
def report():
    c=db(); locs=allowed_locations(c)
    if not locs: c.close(); return page('Report',"<div class='card'>No location assigned.</div>")
    location=request.args.get('location') or locs[0]; location=location if location in locs else locs[0]; year=int(request.args.get('year',date.today().year)); month=max(1,min(12,int(request.args.get('month',date.today().month)))); days=calendar.monthrange(year,month)[1]
    emps=c.execute("SELECT * FROM employees WHERE active=1 AND location=? ORDER BY name",(location,)).fetchall(); rows=''
    for e in emps:
        cnt={s:0 for s in ['P','A','L','WO']}
        for r in c.execute("SELECT status FROM attendance WHERE employee_id=? AND att_date BETWEEN ? AND ?",(e['id'],f'{year:04d}-{month:02d}-01',f'{year:04d}-{month:02d}-{days:02d}')):
            if r['status'] in cnt: cnt[r['status']]+=1
        rows+=f"<tr><td>{e['employee_code']}</td><td>{e['name']}</td><td>{e['designation']}</td><td>{cnt['P']}</td><td>{cnt['A']}</td><td>{cnt['L']}</td><td>{cnt['WO']}</td><td>{sum(cnt.values())}</td></tr>"
    c.close(); return page('Report',f"<div class='hero'><div><h1>Attendance Report</h1><p class='small'>{calendar.month_name[month]} {year} · {location}</p></div><a class='btn success' href='/export?location={location}&year={year}&month={month}'>⬇ Download CSV</a></div><div class='card scroll'><table><tr><th>Employee ID</th><th>Name</th><th>Designation</th><th>Present</th><th>Absent</th><th>Leave</th><th>Weekly Off</th><th>Total</th></tr>{rows}</table></div>")


@app.route('/export')
@login_required
def export():
    c=db(); locs=allowed_locations(c)
    if not locs: c.close(); return 'No location assigned',400
    location=request.args.get('location') or locs[0]; location=location if location in locs else locs[0]; year=int(request.args.get('year',date.today().year)); month=max(1,min(12,int(request.args.get('month',date.today().month)))); days=calendar.monthrange(year,month)[1]
    emps=c.execute("SELECT * FROM employees WHERE active=1 AND location=? ORDER BY name",(location,)).fetchall(); out=io.StringIO(); w=csv.writer(out); w.writerow(['Employee ID','Name','Mobile','Email','Designation','Location']+[str(d) for d in range(1,days+1)]+['Present','Absent','Leave','WO'])
    for e in emps:
        sts=[]; cnt={s:0 for s in ['P','A','L','WO']}
        for d in range(1,days+1):
            r=c.execute("SELECT status FROM attendance WHERE employee_id=? AND att_date=?",(e['id'],f'{year:04d}-{month:02d}-{d:02d}')).fetchone(); s=r['status'] if r else ''; sts.append(s); cnt[s]=cnt.get(s,0)+1 if s in cnt else cnt.get(s,0)
        w.writerow([e['employee_code'],e['name'],e['mobile'],e['email'],e['designation'],e['location']]+sts+[cnt['P'],cnt['A'],cnt['L'],cnt['WO']])
    c.close(); out.seek(0); return send_file(io.BytesIO(out.getvalue().encode('utf-8-sig')),mimetype='text/csv',as_attachment=True,download_name=f'attendance_{year}_{month:02d}.csv')


init()
if __name__=='__main__': app.run(host='0.0.0.0',port=int(os.environ.get('PORT',5000)))
